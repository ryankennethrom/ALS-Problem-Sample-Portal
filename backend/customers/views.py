import csv
import io
from datetime import date, datetime

from django.db import connection, transaction
from django.utils import timezone
from django.db.models import Q
from rapidfuzz.fuzz import WRatio
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status

from .models import Customer, CustomerImport
from .serializers import CustomerSerializer
from .normalization import normalize_header, normalize_customer_type, normalize_company_name, normalize_brand, customer_type_is
from .permissions import IsTrackerAdmin


ALIASES = {
    'external_customer_id': ['coyid', 'customer id', 'customer number', 'account number', 'account', 'id'],
    'company_name': ['company', 'company name', 'customer', 'customer company'],
    'customer_type': ['coytype', 'customer type', 'company type', 'type'],
    'brand': ['brand'],
    'city': ['city'],
    'state': ['state', 'province', 'state/province'],
    'last_date_received': ['lastdaterecd', 'last date recd', 'last date received'],
    'date_created': ['datecreated', 'date created'],
    'primary_contact': ['primarycontact', 'primary contact', 'contact'],
    'email': ['email', 'customer email', 'client contact email', 'email address'],
}


def norm(value):
    return normalize_header(value)


def _normalized_row(row):
    return {norm(k): v for k, v in row.items()}


def value_for(row, field):
    normalized = _normalized_row(row)
    for alias in ALIASES[field]:
        key = norm(alias)
        if key in normalized and normalized[key] not in (None, ''):
            value = normalized[key]
            return value.strip() if isinstance(value, str) else value
    return ''


CUSTOMER_EXPORT_DB_COLUMNS = {
    'external_customer_id',
    'company_name',
    'customer_type',
    'brand',
    'city',
    'state',
    'last_date_received',
    'date_created',
    'primary_contact',
    'email',
}


def missing_customer_schema_columns():
    """Return customer columns that have not been created by migrations yet."""
    table_name = Customer._meta.db_table
    try:
        with connection.cursor() as cursor:
            description = connection.introspection.get_table_description(cursor, table_name)
    except Exception:
        return sorted(CUSTOMER_EXPORT_DB_COLUMNS)
    existing = {column.name for column in description}
    return sorted(CUSTOMER_EXPORT_DB_COLUMNS - existing)


def legacy_customer_unique_constraint_present():
    """Return True while the pre-snapshot (Company, Email) constraint still exists."""
    table_name = Customer._meta.db_table
    try:
        with connection.cursor() as cursor:
            constraints = connection.introspection.get_constraints(cursor, table_name)
    except Exception:
        return False
    return 'unique_customer_company_email' in constraints


def parse_export_date(value):
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value).strip()
    if not raw:
        return None
    for fmt in ('%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%m/%d/%Y', '%m/%d/%y', '%Y/%m/%d'):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            pass
    return None


def _csv_rows(upload):
    # Always rewind because read_export() may inspect the file signature first.
    upload.seek(0)
    raw = upload.read()

    # UTF-8 is preferred, but exports opened/saved through Excel are commonly
    # UTF-16 or Windows-1252. Accept those rather than misreporting a valid
    # customer export as corrupt.
    text = None
    for encoding in ('utf-8-sig', 'utf-16', 'cp1252'):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError('Could not decode the CSV. Save it as UTF-8, UTF-16, or Windows CSV and try again.')
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError('CSV has no header row.')
    return reader.fieldnames, list(reader)


def _xlsx_rows(upload):
    # Always rewind because read_export() may inspect the file signature first.
    upload.seek(0)
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError('Excel import support is not installed on the server.') from exc

    try:
        workbook = load_workbook(upload, read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        header_row = next(iterator, None)
        if not header_row:
            raise ValueError('Excel file has no header row.')
        headers = [str(value or '').strip() for value in header_row]
        rows = []
        for values in iterator:
            rows.append({headers[index]: values[index] if index < len(values) else None for index in range(len(headers))})
        workbook.close()
        return headers, rows
    except ValueError:
        raise
    except Exception as exc:
        raise ValueError('Could not read the Excel workbook.') from exc


def read_export(upload):
    """Read a customer export using both the filename and file signature.

    XLSX files are ZIP containers and begin with the PK signature. Some browser
    or Windows upload paths can provide an unexpected filename, so content
    detection prevents an Excel workbook from ever being sent through the CSV
    decoder (which previously produced the misleading UTF-8 error).
    """
    lower = (getattr(upload, 'name', '') or '').lower()

    upload.seek(0)
    signature = upload.read(8)
    upload.seek(0)

    # Office Open XML workbooks are ZIP containers. Prefer content detection so
    # an actual XLSX is handled as Excel even if its uploaded filename is odd.
    if signature.startswith(b'PK\x03\x04') or lower.endswith('.xlsx'):
        try:
            return _xlsx_rows(upload)
        except ValueError as exc:
            # If the extension says XLSX but the content clearly is not a ZIP,
            # surface an Excel-specific message rather than falling into CSV.
            if lower.endswith('.xlsx'):
                raise
            # A generic ZIP is not a supported customer export.
            raise ValueError('The uploaded file looks like a ZIP archive, not a readable .xlsx workbook.') from exc

    if lower.endswith('.csv'):
        return _csv_rows(upload)

    # If no useful extension was supplied, try text CSV only when the content
    # actually looks textual. This keeps error messages tied to the real format.
    if b'\x00' not in signature:
        try:
            return _csv_rows(upload)
        except ValueError:
            pass

    raise ValueError('Customer export must be a valid .xlsx or .csv file.')


@api_view(['GET'])
@permission_classes([IsTrackerAdmin])
def list_customers(request):
    q = (request.query_params.get('q') or '').strip()
    qs = Customer.objects.all()
    if q:
        qs = qs.filter(
            Q(company_name__icontains=q)
            | Q(email__icontains=q)
            | Q(external_customer_id__icontains=q)
            | Q(primary_contact__icontains=q)
        )
    return Response(CustomerSerializer(qs[:100], many=True).data)


@api_view(['GET'])
def suggest_distributors(request):
    """Fuzzy company-name suggestions restricted to normalized CoyType=Distributor."""
    query = (request.query_params.get('q') or '').strip()
    if not query:
        return Response([])

    # Do not pre-filter CoyType in SQL. Customer exports are external data and can
    # contain harmless whitespace/casing/formatting differences. Normalize the
    # stored value in Python so every row visible in the customer directory gets
    # the same Distributor classification logic. ~20k rows is small enough for
    # this in-memory candidate pass.
    candidates = [
        customer for customer in Customer.objects.exclude(company_name='')
        .only('id', 'external_customer_id', 'company_name', 'customer_type', 'brand', 'city', 'state', 'email')
        if customer_type_is(customer.customer_type, 'Distributor')
    ]

    needle = normalize_company_name(query)
    ranked = []
    for customer in candidates:
        company = (customer.company_name or '').strip()
        haystack = normalize_company_name(company)
        if not haystack:
            continue
        if haystack == needle:
            score = 1000.0
        elif haystack.startswith(needle):
            score = 900.0
        elif needle in haystack:
            score = 800.0
        else:
            score = float(WRatio(needle, haystack))
        if score >= 42:
            ranked.append((score, haystack, customer))

    ranked.sort(key=lambda item: (-item[0], item[1], item[2].id))
    result = []
    seen_companies = set()
    for score, company_key, customer in ranked:
        if company_key in seen_companies:
            continue
        seen_companies.add(company_key)
        data = CustomerSerializer(customer).data
        data['match_score'] = round(score, 2)
        result.append(data)
        if len(result) >= 15:
            break
    return Response(result)


@api_view(['GET'])
def suggest_end_users(request):
    """Fuzzy company-name suggestions restricted to CoyType=End User."""
    query = (request.query_params.get('q') or '').strip()
    if not query:
        return Response([])

    candidates = [
        customer for customer in Customer.objects.exclude(company_name='')
        .only('id', 'external_customer_id', 'company_name', 'customer_type', 'brand', 'city', 'state', 'email')
        if customer_type_is(customer.customer_type, 'End User')
    ]
    needle = normalize_company_name(query)
    ranked = []
    for customer in candidates:
        company = (customer.company_name or '').strip()
        if not company:
            continue
        haystack = normalize_company_name(company)
        if haystack == needle:
            score = 1000.0
        elif haystack.startswith(needle):
            score = 900.0
        elif needle in haystack:
            score = 800.0
        else:
            score = float(WRatio(needle, haystack))
        if score >= 42:
            ranked.append((score, company.casefold(), customer))

    ranked.sort(key=lambda item: (-item[0], item[1]))
    result = []
    seen_companies = set()
    for score, _name, customer in ranked:
        company_key = (customer.company_name or '').strip().casefold()
        if not company_key or company_key in seen_companies:
            continue
        seen_companies.add(company_key)
        data = CustomerSerializer(customer).data
        data['match_score'] = round(score, 2)
        result.append(data)
        if len(result) >= 15:
            break
    return Response(result)


@api_view(['GET'])
def suggest_brands(request):
    """Fuzzy suggestions from distinct Brand values in the current Customer Export."""
    query = (request.query_params.get('q') or '').strip()
    if not query:
        return Response([])

    # Preserve the source spelling/capitalization while deduplicating harmless
    # whitespace and case differences from the imported customer directory.
    unique = {}
    counts = {}
    examples = {}
    for customer in Customer.objects.exclude(brand='').only('brand', 'company_name'):
        brand = (customer.brand or '').strip()
        key = normalize_brand(brand)
        if not key:
            continue
        unique.setdefault(key, brand)
        counts[key] = counts.get(key, 0) + 1
        company = (customer.company_name or '').strip()
        if company:
            bucket = examples.setdefault(key, [])
            company_key = normalize_company_name(company)
            if company_key and all(normalize_company_name(item) != company_key for item in bucket) and len(bucket) < 3:
                bucket.append(company)

    needle = normalize_brand(query)
    ranked = []
    for key, brand in unique.items():
        if key == needle:
            score = 1000.0
        elif key.startswith(needle):
            score = 900.0
        elif needle in key:
            score = 800.0
        else:
            score = float(WRatio(needle, key))
        if score >= 42:
            ranked.append((score, key, brand))

    ranked.sort(key=lambda item: (-item[0], item[1]))
    return Response([
        {
            'brand': brand,
            'match_score': round(score, 2),
            'customer_count': counts.get(key, 0),
            'company_examples': examples.get(key, []),
        }
        for score, key, brand in ranked[:20]
    ])


@api_view(['GET'])
def suggest_client_emails(request):
    """Fuzzy email/contact suggestions using prioritized company fallbacks."""
    query = (request.query_params.get('q') or '').strip()
    companies = []
    seen_companies = set()
    for raw_company in request.query_params.getlist('company'):
        company = (raw_company or '').strip()
        key = company.casefold()
        if company and key not in seen_companies:
            companies.append(company)
            seen_companies.add(key)

    # With configured dependencies, choose the first populated company that has
    # at least one imported email. This choice is independent of the user's fuzzy
    # query, so typing narrows within one active company rather than unexpectedly
    # switching to a lower-priority fallback.
    active_company = ''
    active_priority = None
    queryset = Customer.objects.exclude(email='')
    if companies:
        scoped = None
        for index, company in enumerate(companies):
            candidate = Customer.objects.exclude(email='').filter(company_name__iexact=company)
            if candidate.exists():
                active_company = company
                active_priority = index
                scoped = candidate
                break
        if scoped is None:
            return Response({
                'results': [],
                'active_company': '',
                'active_priority': None,
                'attempted_companies': companies,
            })
        queryset = scoped
    elif not query:
        # An unscoped Client Email field should not dump the entire directory on focus.
        return Response({
            'results': [],
            'active_company': '',
            'active_priority': None,
            'attempted_companies': [],
        })

    candidates = list(queryset.only(
        'id', 'external_customer_id', 'company_name', 'customer_type', 'brand',
        'city', 'state', 'primary_contact', 'email',
    ))
    needle = query.casefold()
    ranked = []
    for customer in candidates:
        email = (customer.email or '').strip()
        if not email:
            continue
        contact = (customer.primary_contact or '').strip()
        company_name = (customer.company_name or '').strip()

        if not query:
            score = 100.0
        else:
            email_folded = email.casefold()
            if email_folded == needle:
                score = 220.0
            elif email_folded.startswith(needle):
                score = 195.0
            elif needle in email_folded:
                score = 175.0
            else:
                score = float(WRatio(query, email))
            if contact:
                score = max(score, float(WRatio(query, contact)) * 0.9)
            if company_name:
                score = max(score, float(WRatio(query, company_name)) * 0.65)

        if score >= 38:
            ranked.append((score, email.casefold(), customer))

    ranked.sort(key=lambda item: (-item[0], item[1]))
    seen = set()
    result = []
    for score, _email, customer in ranked:
        key = ((customer.email or '').casefold(), (customer.company_name or '').casefold())
        if key in seen:
            continue
        seen.add(key)
        data = CustomerSerializer(customer).data
        data['match_score'] = round(score, 2)
        result.append(data)
        # Scoped empty-query mode intentionally returns every unique email under
        # the active company. Typed fuzzy searches stay compact for usability.
        if query and len(result) >= 30:
            break

    return Response({
        'results': result,
        'active_company': active_company,
        'active_priority': active_priority,
        'attempted_companies': companies,
    })


def _customer_import_payload(batch):
    uploader = batch.imported_by
    uploader_name = 'Unknown user'
    uploader_username = ''
    if uploader is not None:
        uploader_name = uploader.get_full_name().strip() or uploader.username
        uploader_username = uploader.username

    imported_local_date = timezone.localdate(batch.imported_at)
    days_ago = max((timezone.localdate() - imported_local_date).days, 0)
    return {
        'id': batch.id,
        'filename': batch.filename,
        'imported_at': batch.imported_at,
        'row_count': batch.row_count,
        'days_ago': days_ago,
        'uploaded_by': {
            'id': uploader.id if uploader is not None else None,
            'username': uploader_username,
            'name': uploader_name,
        },
    }


@api_view(['GET'])
@permission_classes([IsTrackerAdmin])
def customer_overview(request):
    history_qs = (
        CustomerImport.objects.select_related('imported_by')
        .order_by('-imported_at', '-id')
    )
    history = [_customer_import_payload(batch) for batch in history_qs[:100]]
    return Response({
        'row_count': Customer.objects.count(),
        'history_count': history_qs.count(),
        'latest_upload': history[0] if history else None,
        'history': history,
    })


@api_view(['POST'])
@permission_classes([IsTrackerAdmin])
def import_customers(request):
    upload = request.FILES.get('file')
    if not upload:
        return Response({'detail': 'Attach a .xlsx or .csv file using the field name `file`.'}, status=status.HTTP_400_BAD_REQUEST)

    missing_schema = missing_customer_schema_columns()
    if missing_schema:
        return Response(
            {
                'detail': (
                    'Customer database schema is out of date. Run `python manage.py migrate` '
                    f'and retry the import. Missing database column(s): {", ".join(missing_schema)}.'
                )
            },
            status=status.HTTP_503_SERVICE_UNAVAILABLE,
        )

    if legacy_customer_unique_constraint_present():
        return Response(
            {
                'detail': (
                    'Customer database schema is out of date. Run `python manage.py migrate` '
                    'and retry the import. The old Company + Email uniqueness constraint still exists.'
                )
            },
            status=status.HTTP_503_SERVICE_UNAVAILABLE,
        )

    try:
        fieldnames, rows = read_export(upload)
    except ValueError as exc:
        return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    normalized_headers = {norm(name) for name in fieldnames}
    required = {'coyid', 'company'}
    missing = sorted(required - normalized_headers)
    if missing:
        return Response(
            {'detail': f'Customer export is missing required column(s): {", ".join(missing)}.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # The Customer Export is a snapshot, not an upsert feed. None of its columns
    # are guaranteed to be unique, including CoyId, Company, or Email. Parse the
    # complete snapshot first, then atomically replace the previous directory.
    # This preserves duplicate source rows while also preventing repeated uploads
    # from accumulating another ~20k duplicate records.
    customer_rows = []
    for row in rows:
        values = {
            'external_customer_id': str(value_for(row, 'external_customer_id') or '').strip(),
            'company_name': str(value_for(row, 'company_name') or '').strip(),
            'customer_type': normalize_customer_type(value_for(row, 'customer_type')),
            'brand': str(value_for(row, 'brand') or '').strip(),
            'city': str(value_for(row, 'city') or '').strip(),
            'state': str(value_for(row, 'state') or '').strip(),
            'last_date_received': parse_export_date(value_for(row, 'last_date_received')),
            'date_created': parse_export_date(value_for(row, 'date_created')),
            'primary_contact': str(value_for(row, 'primary_contact') or '').strip(),
            'email': str(value_for(row, 'email') or '').strip().lower(),
        }

        # Ignore only completely empty spreadsheet rows. Do not use any customer
        # field as a uniqueness key or deduplicate rows from the source export.
        if not any(value not in ('', None) for value in values.values()):
            continue
        customer_rows.append(values)

    # Perform the replacement only after the entire upload has been parsed and
    # normalized successfully. The transaction makes this behave like a swap:
    # readers see the old directory until the replacement commits, and any
    # insertion failure restores the previous customer directory automatically.
    with transaction.atomic():
        previous_count = Customer.objects.count()
        previous_import_count = CustomerImport.objects.count()

        # The current Customer Export is the whole source of truth for customer
        # rows, but import records are an audit trail and must be retained.
        Customer.objects.all().delete()

        batch = CustomerImport.objects.create(
            filename=upload.name,
            imported_by=request.user,
            row_count=len(customer_rows),
        )
        Customer.objects.bulk_create(
            [Customer(source_import=batch, **values) for values in customer_rows],
            batch_size=1000,
        )

    return Response(
        {
            'imported': len(customer_rows),
            'source_rows': len(rows),
            'replaced': previous_count,
            'previous_imports': previous_import_count,
            'import_history_count': previous_import_count + 1,
            'filename': upload.name,
            'mode': 'replace',
        },
        status=status.HTTP_201_CREATED,
    )

